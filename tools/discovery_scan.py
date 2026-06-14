#!/usr/bin/env -S uv run
# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "pyserial",
#   "openpyxl",
#   "Pillow",
# ]
# ///
"""
discovery_scan.py — BondTest72 discovery scan tool

Sends DISCOVERY_SCAN to the tester, collects sense voltages for every mez-pin
pair (70×70), then writes three outputs:
  • .csv   — raw voltage matrix indexed by die pad (0-based)
  • .xlsx  — same with color-scale conditional formatting
  • .png   — pixel heatmap (green=conducting, red=isolated)

Rows = source die pad (current injected via Bus::D)
Cols = sink die pad   (tester GND return via Bus::B)
Cell = sense voltage at Bus::D (ADC0) in volts, 3 d.p.

Usage:
    uv run tools/discovery_scan.py --port /dev/tty.usbmodem1101 --padmap 2
    uv run tools/discovery_scan.py --port COM3 --padmap 3 --out my_die_scan --cell 8
"""

import argparse
import colorsys
import csv
import sys
import time
from datetime import datetime

try:
    import serial
except ImportError:
    print("ERROR: pyserial not installed. Run: pip install pyserial")
    sys.exit(1)

# ── Pad map definitions ────────────────────────────────────────────────────────
#
# All die pads are 0-based (matching firmware pad_map_registry.cpp).
# mez_to_die: mez (adapter) pin → die pad.
# nc_die_pads: die pads physically present on die but with no mez connection.
# all_die_count: total die pads (0 .. all_die_count-1).

PAD_MAP_DEFS = {
    1: {
        "name": "1x1 Mezzanine70 v1",
        "mez_to_die": {
            # die pads 0–7 (mez 63–70)
            63:  0,  64:  1,  65:  2,  66:  3,  67:  4,  68:  5,  69:  6,  70:  7,
            # gap: GND die pad 8 (no mez connection)
            # die pads 9–16 (mez 1–8)
             1:  9,   2: 10,   3: 11,   4: 12,   5: 13,   6: 14,   7: 15,   8: 16,
            # die pad 17 VDDIO (mez 9), die pad 18 GND (mez 10)
             9: 17,  10: 18,
            # die pads 19–24 (mez 11–16)
            11: 19,  12: 20,  13: 21,  14: 22,  15: 23,  16: 24,
            # die pad 25 PWR_AUX (mez 17), die pad 26 GND (mez 18)
            17: 25,  18: 26,
            # die pads 27–32 (mez 19–24)
            19: 27,  20: 28,  21: 29,  22: 30,  23: 31,  24: 32,
            # gap: die pad 33 N/C
            # die pad 34 VDD_CORE (mez 25), die pad 35 GND (mez 26), die pad 36 VDDIO (mez 27)
            25: 34,  26: 35,  27: 36,
            # die pads 37–44 (mez 28–35); apin34/die43 IS connected on v1
            28: 37,  29: 38,  30: 39,  31: 40,  32: 41,  33: 42,  34: 43,  35: 44,
            # gap: die pad 45 N/C
            # die pads 46–55 (mez 36–45)
            36: 46,  37: 47,  38: 48,  39: 49,  40: 50,  41: 51,  42: 52,  43: 53,  44: 54,  45: 55,
            # die pad 56 GND (mez 46), die pad 57 VDDIO (mez 47)
            46: 56,  47: 57,
            # die pads 58–61 (mez 48–51)
            48: 58,  49: 59,  50: 60,  51: 61,
            # gap: die pad 62 GND (mez 53), die pad 63 N/C in v1 (mez 52 trace unconnected)
            53: 62,
            # die pads 64–69 (mez 54–59)
            54: 64,  55: 65,  56: 66,  57: 67,  58: 68,  59: 69,
            # gap: die pad 70 N/C
            # die pad 71 VDD_CORE (mez 60), die pad 72 GND (mez 61), die pad 73 VDDIO (mez 62)
            60: 71,  61: 72,  62: 73,
        },
        "all_die_count": 74,   # die pads 0..73
        "nc_die_pads":   {8, 33, 45, 63, 70},  # GND or unconnected — no mez pin
    },
    2: {
        "name": "1x1 Mezzanine70 v2",
        "mez_to_die": {
            63:  0,  64:  1,  65:  2,  66:  3,  67:  4,  68:  5,  69:  6,  70:  7,
             1:  9,   2: 10,   3: 11,   4: 12,   5: 13,   6: 14,   7: 15,   8: 16,
             9: 17,  10: 18,
            11: 19,  12: 20,  13: 21,  14: 22,  15: 23,  16: 24,
            17: 25,  18: 26,
            19: 27,  20: 28,  21: 29,  22: 30,  23: 31,  24: 32,
            25: 34,  26: 35,  27: 36,
            28: 37,  29: 38,  30: 39,  31: 40,  32: 41,  33: 42,  34: 43,  35: 44,
            36: 46,  37: 47,  38: 48,  39: 49,  40: 50,  41: 51,  42: 52,  43: 53,  44: 54,  45: 55,
            46: 56,  47: 57,
            48: 58,  49: 59,  50: 60,  51: 61,
            # die pad 62 GND (mez 53), die pad 63 PWR_AUX (mez 52) — connected in v2
            52: 63,  53: 62,
            54: 64,  55: 65,  56: 66,  57: 67,  58: 68,  59: 69,
            60: 71,  61: 72,  62: 73,
        },
        "all_die_count": 74,
        "nc_die_pads":   {8, 33, 45, 70},
    },
    3: {
        "name": "1x0p5 Mezzanine70 v1",
        "mez_to_die": {
            # die pads 0–3 (mez 63–66): clk, rst_n, bidir_0–1
            63:  0,  64:  1,  65:  2,  66:  3,
            # die pad 4 GND (mez 61), die pad 5 VDD_CORE 1 (mez 60)
            61:  4,  60:  5,
            # die pads 6–9 (mez 67–70): bidir_2–5
            67:  6,  68:  7,  69:  8,  70:  9,
            # die pads 10–17 (mez 1–8): bidir_6–13
             1: 10,   2: 11,   3: 12,   4: 13,   5: 14,   6: 15,   7: 16,   8: 17,
            # die pad 18 GND (mez 10), die pad 19 VDDIO 0 (mez 9)
            10: 18,   9: 19,
            # die pads 20–23 (mez 11–14): bidir_14–17
            11: 20,  12: 21,  13: 22,  14: 23,
            # die pad 24 GND (mez 18), die pad 25 PWR_AUX 0 (mez 17)
            18: 24,  17: 25,
            # die pads 26–33 (mez 15–16, 19–24): bidir_18–25
            15: 26,  16: 27,  19: 28,  20: 29,  21: 30,  22: 31,  23: 32,  24: 33,
            # gap: die pad 34 N/C (GND on die, no mez pin)
            # die pad 35 VDDIO 1 (mez 27)
            27: 35,
            # die pads 36–39 (mez 28–31): bidir_26–29
            28: 36,  29: 37,  30: 38,  31: 39,
            # die pad 40 VDD_CORE (mez 25), die pad 41 GND (mez 26)
            25: 40,  26: 41,
            # die pads 42–53 (mez 32–43): bidir_30–41
            32: 42,  33: 43,  34: 44,  35: 45,  36: 46,  37: 47,
            38: 48,  39: 49,  40: 50,  41: 51,  42: 52,  43: 53,
            # die pad 54 VDDIO 2 (mez 47), die pad 55 GND (mez 46)
            47: 54,  46: 55,
            # die pads 56–59 (mez 44–45, 48–49): bidir_42–45
            44: 56,  45: 57,  48: 58,  49: 59,
            # die pad 60 PWR_AUX 1 (mez 52), die pad 61 GND (mez 53)
            52: 60,  53: 61,
            # die pads 62–69 (mez 50–51, 54–55, 59–56): an_0–3, in_0–3
            50: 62,  51: 63,  54: 64,  55: 65,  59: 66,  58: 67,  57: 68,  56: 69,
            # die pad 70 VDDIO 3 (mez 62); gap: die pad 71 N/C (GND on die, outside range)
            62: 70,
        },
        "all_die_count": 72,   # die pads 0..71
        "nc_die_pads":   {34, 71},  # die34: GND on die, no mez pin; die71: GND on die, no mez pin
    },
}


def load_padmap(pm_id: int) -> tuple:
    """Return (mez_to_die, die_to_mez, nc_die_pads, all_die_pads, name)."""
    d = PAD_MAP_DEFS[pm_id]
    mez_to_die  = d["mez_to_die"]
    die_to_mez  = {v: k for k, v in mez_to_die.items()}
    nc_die_pads = d["nc_die_pads"]
    all_die_pads = list(range(d["all_die_count"]))
    return mez_to_die, die_to_mez, nc_die_pads, all_die_pads, d["name"]


def label_indices(all_die_pads: list) -> set:
    n = len(all_die_pads)
    return {i for i in range(n) if (all_die_pads[i] % 10 == 0) or i == 0 or i == n - 1}


def col_label(die: int) -> str:
    return str(die)


# ── Serial scan ───────────────────────────────────────────────────────────────

def run_scan(port: str, baud: int, timeout: int) -> dict:
    """Return dict keyed (src_mez, snk_mez) → voltage float."""
    print(f"Opening {port} at {baud} baud …")
    ser = serial.Serial(port, baud, timeout=2)
    time.sleep(0.5)
    ser.reset_input_buffer()

    print("Sending DISCOVERY_SCAN …")
    ser.write(b"DISCOVERY_SCAN\n")
    ser.flush()

    data = {}
    start = time.time()
    total_expected = 70 * 69
    received = 0
    last_progress = -1

    while True:
        if time.time() - start > timeout:
            print(f"\nERROR: timed out after {timeout}s ({received}/{total_expected} received)")
            ser.close()
            sys.exit(1)

        line = ser.readline().decode("ascii", errors="replace").strip()
        if not line:
            continue

        if line == "DSCAN DONE":
            print(f"\rReceived {received}/{total_expected} measurements. Done.    ")
            break

        if line.startswith("DSCAN "):
            parts = line.split()
            if len(parts) >= 4:
                if '=' in parts[1]:
                    d = dict(p.split('=', 1) for p in parts[1:] if '=' in p)
                    src, snk, v = int(d.get('src', '0')), int(d.get('snk', '0')), float(d.get('sv', '0'))
                else:
                    src, snk, v = int(parts[1]), int(parts[2]), float(parts[3])
                data[(src, snk)] = v
                received += 1
                pct = received * 100 // total_expected
                if pct != last_progress:
                    print(f"\r  {pct:3d}%  ({received}/{total_expected})", end="", flush=True)
                    last_progress = pct
        elif line.startswith("ERROR "):
            rest = line[6:]
            if '=' in rest:
                kv = dict(p.split('=', 1) for p in rest.split() if '=' in p)
                print(f"\nERROR from tester (code {kv.get('code', '?')}): {kv.get('msg', rest)}")
            else:
                print(f"\nERROR from tester: {line}")
            ser.close()
            sys.exit(1)
        else:
            print(f"\n[tester] {line}", end="")

    ser.close()
    return data


# ── Matrix builder ────────────────────────────────────────────────────────────

def build_matrix(data: dict, die_to_mez: dict, nc_die_pads: set, all_die_pads: list) -> list[list]:
    """Build NxN matrix (list of rows, each a list of cell strings)."""
    rows = []
    for src_die in all_die_pads:
        row = []
        for snk_die in all_die_pads:
            if src_die == snk_die:
                row.append("")
                continue
            if src_die in nc_die_pads or snk_die in nc_die_pads:
                row.append("N/C")
                continue
            src_mez = die_to_mez.get(src_die)
            snk_mez = die_to_mez.get(snk_die)
            if src_mez is None or snk_mez is None:
                row.append("N/C")
                continue
            v = data.get((src_mez, snk_mez))
            row.append(f"{v:.3f}" if v is not None else "")
        rows.append(row)
    return rows


# ── CSV output ────────────────────────────────────────────────────────────────

def write_csv(matrix: list[list], all_die_pads: list, path: str):
    header = ["src \\ snk (die pad)"] + [col_label(d) for d in all_die_pads]
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        for i, row in enumerate(matrix):
            w.writerow([col_label(all_die_pads[i])] + row)
    print(f"CSV written → {path}")


# ── Excel output ──────────────────────────────────────────────────────────────

def write_excel(matrix: list[list], all_die_pads: list, path: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.formatting.rule import ColorScaleRule
    except ImportError:
        print("openpyxl not installed — skipping Excel output. Run: pip install openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Discovery Scan"

    header_fill = PatternFill("solid", fgColor="D0D0D0")
    header_font = Font(bold=True, size=8)
    cell_font   = Font(size=8)
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal="center", vertical="center")
    labels      = [col_label(d) for d in all_die_pads]
    n           = len(all_die_pads)

    ws.cell(1, 1, "die pad").font = header_font
    ws.cell(1, 1).fill = header_fill
    ws.cell(1, 1).alignment = center

    for col_idx, lbl in enumerate(labels, start=2):
        c = ws.cell(1, col_idx, lbl)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[c.column_letter].width = 3

    for row_idx, (src_die, row) in enumerate(zip(all_die_pads, matrix), start=2):
        lc = ws.cell(row_idx, 1, labels[row_idx - 2])
        lc.font = header_font
        lc.fill = header_fill
        lc.alignment = center
        lc.border = border

        for col_idx, val in enumerate(row, start=2):
            c = ws.cell(row_idx, col_idx, val)
            c.font = cell_font
            c.alignment = center
            c.border = border
            if val and val != "N/C":
                try:
                    c.value = float(val)
                except ValueError:
                    pass

    data_range = f"B2:{ws.cell(n + 1, n + 1).coordinate}"
    ws.conditional_formatting.add(
        data_range,
        ColorScaleRule(
            start_type="num", start_value=0.0,  start_color="63BE7B",
            mid_type="num",   mid_value=1.65,   mid_color="FFEB84",
            end_type="num",   end_value=3.3,    end_color="F8696B",
        )
    )

    ws.row_dimensions[1].height = 40
    ws.column_dimensions["A"].width = 4
    ws.freeze_panes = "B2"

    wb.save(path)
    print(f"Excel written → {path}")


# ── PNG heatmap ───────────────────────────────────────────────────────────────

def _voltage_to_color(val: str, is_diag: bool) -> tuple:
    if is_diag:
        return (200, 200, 200)
    if not val or val == "":
        return (220, 220, 220)
    if val == "N/C":
        return (210, 210, 210)
    try:
        v = float(val)
    except ValueError:
        return (70, 70, 70)
    v = max(0.0, min(3.3, v))
    hue = (1.0 - v / 3.3) * 120.0 / 360.0
    r, g, b = colorsys.hsv_to_rgb(hue, 0.85, 0.88)
    return (int(r * 255), int(g * 255), int(b * 255))


def write_image(matrix: list[list], all_die_pads: list, path: str, cell: int = 5):
    try:
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        print("Pillow not installed — skipping PNG output. Run: pip install Pillow")
        return

    MARGIN = 26
    N      = len(all_die_pads)
    BG     = (255, 255, 255)
    lbl_idx = label_indices(all_die_pads)

    img  = Image.new("RGB", (MARGIN + N * cell, MARGIN + N * cell), BG)
    draw = ImageDraw.Draw(img)
    font = ImageFont.load_default()

    for ri, row in enumerate(matrix):
        for ci, val in enumerate(row):
            x = MARGIN + ci * cell
            y = MARGIN + ri * cell
            draw.rectangle([x, y, x + cell - 1, y + cell - 1],
                           fill=_voltage_to_color(val, ri == ci))

    for i, die in enumerate(all_die_pads):
        if i in lbl_idx:
            y = MARGIN + i * cell + cell // 2 - 4
            draw.text((1, y), str(die), fill=(80, 80, 80), font=font)

    for i, die in enumerate(all_die_pads):
        if i in lbl_idx:
            x     = MARGIN + i * cell + cell // 2
            label = str(die)
            tw    = len(label) * 6
            tmp   = Image.new("RGB", (tw + 2, 10), BG)
            ImageDraw.Draw(tmp).text((0, 0), label, fill=(80, 80, 80), font=font)
            tmp   = tmp.rotate(90, expand=True)
            img.paste(tmp, (x - 4, 1))

    img.save(path)
    print(f"PNG written  → {path}")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description="BondTest72 discovery scan — sweeps all pad pairs and saves voltage matrix"
    )
    ap.add_argument("--port",    required=True,  help="Serial port (e.g. /dev/tty.usbmodem1101 or COM3)")
    ap.add_argument("--padmap", required=True, type=int, choices=[1, 2, 3],
                    help="Pad map ID: 1=1x1 v1, 2=1x1 v2, 3=1x0p5 v1")
    ap.add_argument("--baud",    type=int, default=115200, help="Baud rate (default: 115200)")
    ap.add_argument("--out",     default=None,             help="Output basename without extension")
    ap.add_argument("--name",    default=None,             help="Label appended to auto-generated filename")
    ap.add_argument("--timeout", type=int, default=120,    help="Scan timeout in seconds (default: 120)")
    ap.add_argument("--cell",    type=int, default=16,     help="Heatmap cell size in pixels (default: 16)")
    args = ap.parse_args()

    mez_to_die, die_to_mez, nc_die_pads, all_die_pads, pm_name = load_padmap(args.padmap)
    print(f"Pad map {args.padmap}: {pm_name}  ({len(all_die_pads)} die pads, {len(nc_die_pads)} N/C)")

    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix   = f"_{args.name}" if args.name else ""
    out_base = args.out or f"discovery_pm{args.padmap}_{ts}{suffix}"

    data   = run_scan(args.port, args.baud, args.timeout)
    matrix = build_matrix(data, die_to_mez, nc_die_pads, all_die_pads)

    write_csv(matrix,   all_die_pads, out_base + ".csv")
    write_excel(matrix, all_die_pads, out_base + ".xlsx")
    write_image(matrix, all_die_pads, out_base + ".png", args.cell)

    print(f"\nDone. {len(data)} / {70 * 69} measurements.")


if __name__ == "__main__":
    main()
